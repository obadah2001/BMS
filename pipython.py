import time #my code
import paho.mqtt.client as mqtt #my code
import ssl #my code
import json #my code
import _thread #my code
import RPi.GPIO as GPIO #my code
import board
import adafruit_dht
import serial
import atexit
from openpyxl import load_workbook

#tkinter
from tkinter import *
from datetime import datetime
import tkinter as tk
import adafruit_dht as dht
import threading
import tkinter.font as TkFont
from PIL import ImageTk,Image
import RPi.GPIO as GPIO

ser = serial.Serial('/dev/ttyACM0', 9600)
client = mqtt.Client() # my code

vMax = 13.66
vMin = 12
switch = 1

def init():
    client.on_connect = on_connect # my code
    client.tls_set(ca_certs='./rootCA.pem', certfile='./certificate.pem.crt', keyfile='./private.pem.key', tls_version=ssl.PROTOCOL_SSLv23) # my code
    client.tls_insecure_set(True) # my code
    client.connect("aa7bx2yvfl8gx-ats.iot.us-east-1.amazonaws.com", 8883, 60) #Taken from REST API endpoint - Use your own. 

def onExit():
    print("bye")
    
def jsonConstruct(input):
    data = input.split()
    voltage = data[0]
    current = data[1]
    try:
        temperature = data[2]
    except:
        temperature = 0
    try:
        humidity = data[3]
    except:
        humidity = 0
    
               ####Calculations###
    #Fahrenheit
    temperatureF = float(temperature) * (9.0/5.0) + 32.0
    #Power
    power = float(voltage) * float(current)
    
    
    output  = "{\"CelsiusTemperature\":"
    output += temperature
    output += ",\"FahrenheitTemperature\":"
    output += str(temperatureF)
    output += ",\"Humidity\":"
    output += humidity
    output += ",\"Voltage\":"
    output += voltage
    output += ",\"Current\":"
    output += current
    output += ",\"Power\":"
    output += str(power)
    output += "}"
    return output

def calculateBatteryPercent(voltage):
    #this is just simple battery percentage calculation
    numCells = 193
    for i in range(1,37):
        if sheet.cell(row=i, column=2).value == voltage:
            return (numCells-sheet.cell(row=i, column=1).value)*(100/numCells)
    return 0
    
    

def on_connect(client, userdata, flags, rc): #my code
    print("Connected with result code "+str(rc)) #my code

#init()
#atexit.register(onExit)
workbook = load_workbook(filename="data.xlsx")
#dataLog = load_workbook(filename="datalog.xlsx")
sheet = workbook.active
f = open("timemarks.txt","a")
f.write(datetime.now().strftime("%m/%d/%Y, %H:%M:%S"))
f.close()
while True:  #my code
    data = ser.readline().decode().strip()
    jsonStr = jsonConstruct(data)
    print(jsonStr)
    jsonObj = json.loads(jsonStr)
    f = open("data.txt","a")
    f.write(jsonStr)
    f.write("\n")
    f.close()
    client.publish("device/data", payload=(jsonStr) , qos=0, retain=False) #my code
    time.sleep(5) #my code
client.loop_forever() #my code


